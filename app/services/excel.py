"""Layanan generate file Excel (.xlsx) untuk seluruh laporan.

Menggunakan openpyxl. Setiap fungsi `build_*` mengembalikan bytes (.xlsx).
Format umum:
  Baris 1 : Nama perusahaan (kop)
  Baris 2 : Judul laporan
  Baris 3 : Periode
  Baris 4 : kosong
  Baris 5+: header / data
"""
import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style presets ────────────────────────────────────────────────────────────
_THIN  = Side(style="thin",  color="999999")
_MED   = Side(style="medium", color="111111")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_HEADER  = PatternFill("solid", fgColor="E2E2E2")
_FILL_SECTION = PatternFill("solid", fgColor="F4F4F4")
_FILL_TOTAL   = PatternFill("solid", fgColor="EBEBEB")

_FONT_TITLE   = Font(name="Arial", size=14, bold=True)
_FONT_SUBTTL  = Font(name="Arial", size=12, bold=True)
_FONT_PERIOD  = Font(name="Arial", size=10, italic=True, color="555555")
_FONT_HEADER  = Font(name="Arial", size=10, bold=True)
_FONT_BOLD    = Font(name="Arial", size=10, bold=True)
_FONT_NORMAL  = Font(name="Arial", size=10)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

# Format angka rupiah (tanpa simbol Rp, pakai pemisah ribuan & tanda kurung untuk negatif)
_NUM_FMT = '#,##0;(#,##0);"-"'


# ── Helpers ──────────────────────────────────────────────────────────────────
def _letterhead(ws, company: str, title: str, period: str, n_cols: int) -> int:
    """Tulis kop (3 baris) + 1 baris kosong. Return baris berikutnya untuk data."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=company).font = _FONT_TITLE
    ws.cell(row=1, column=1).alignment = _ALIGN_CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=title).font = _FONT_SUBTTL
    ws.cell(row=2, column=1).alignment = _ALIGN_CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws.cell(row=3, column=1, value=period).font = _FONT_PERIOD
    ws.cell(row=3, column=1).alignment = _ALIGN_CENTER

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    return 5  # baris ke-4 dibiarkan kosong


def _set_widths(ws, widths: Iterable[float]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, row: int, headers: list[str]):
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _FONT_HEADER
        c.alignment = _ALIGN_CENTER
        c.fill = _FILL_HEADER
        c.border = _BORDER_THIN
    ws.row_dimensions[row].height = 22


def _as_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1. Jurnal Umum ───────────────────────────────────────────────────────────
def build_jurnal_umum(company, period_str, rows, grand_total) -> bytes:
    """rows: list of {tanggal, keterangan, kode, akun, is_kredit, debet, kredit}.

    Catatan: kolom 'Kode' tidak ditampilkan; nomor akun dipindahkan ke kolom Ref.
    """
    wb = Workbook(); ws = wb.active; ws.title = "Jurnal Umum"
    _set_widths(ws, [10, 35, 30, 10, 16, 16])
    r = _letterhead(ws, company, "Jurnal Umum", period_str, 6)

    _write_header(ws, r, ["Tanggal", "Nama Akun", "Keterangan", "Ref", "Debet", "Kredit"])
    r += 1

    for row in rows:
        ws.cell(row=r, column=1, value=row.get("tanggal", ""))
        akun_cell = ws.cell(row=r, column=2, value=row.get("akun", ""))
        if row.get("is_kredit"):
            akun_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.cell(row=r, column=3, value=row.get("keterangan", ""))
        # Ref diisi dengan kode akun (bukan "JU")
        ws.cell(row=r, column=4, value=row.get("kode", "")).alignment = _ALIGN_CENTER
        d = ws.cell(row=r, column=5, value=row["debet"]  if row["debet"]  else None)
        k = ws.cell(row=r, column=6, value=row["kredit"] if row["kredit"] else None)
        d.number_format = k.number_format = _NUM_FMT
        d.alignment = k.alignment = _ALIGN_RIGHT
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _BORDER_THIN
        r += 1

    # Total
    ws.cell(row=r, column=1, value="Total").font = _FONT_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for col, val in [(5, grand_total), (6, grand_total)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)
        if cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = _FILL_TOTAL

    return _as_bytes(wb)


# ── 1b. Jurnal Penyesuaian ───────────────────────────────────────────────────
def build_jurnal_penyesuaian(company, period_str, rows, grand_total) -> bytes:
    """Sama seperti Jurnal Umum tapi TANPA kolom Ref."""
    wb = Workbook(); ws = wb.active; ws.title = "Jurnal Penyesuaian"
    _set_widths(ws, [10, 38, 32, 16, 16])
    r = _letterhead(ws, company, "Jurnal Penyesuaian", period_str, 5)

    _write_header(ws, r, ["Tanggal", "Nama Akun", "Keterangan", "Debet", "Kredit"])
    r += 1

    for row in rows:
        ws.cell(row=r, column=1, value=row.get("tanggal", ""))
        akun_cell = ws.cell(row=r, column=2, value=row.get("akun", ""))
        if row.get("is_kredit"):
            akun_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.cell(row=r, column=3, value=row.get("keterangan", ""))
        d = ws.cell(row=r, column=4, value=row["debet"]  if row["debet"]  else None)
        k = ws.cell(row=r, column=5, value=row["kredit"] if row["kredit"] else None)
        d.number_format = k.number_format = _NUM_FMT
        d.alignment = k.alignment = _ALIGN_RIGHT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = _BORDER_THIN
        r += 1

    # Total
    ws.cell(row=r, column=1, value="Total").font = _FONT_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for col, val in [(4, grand_total), (5, grand_total)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)
        if cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = _FILL_TOTAL

    return _as_bytes(wb)


# ── 2. Buku Besar ────────────────────────────────────────────────────────────
def build_buku_besar(company, period_str, ledger) -> bytes:
    """ledger: list of {akun, saldo_akhir, rows[{tanggal, keterangan, ref, debet, kredit, saldo}]}"""
    wb = Workbook(); ws = wb.active; ws.title = "Buku Besar"
    _set_widths(ws, [12, 35, 8, 16, 16, 18])
    r = _letterhead(ws, company, "Buku Besar", period_str, 6)

    for item in ledger:
        akun = item["akun"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=f"{akun.kode_akun}  {akun.nama_akun}  ({akun.jenis_akun}, saldo normal: {akun.saldo_normal})")
        c.font = _FONT_SUBTTL
        c.fill = _FILL_SECTION
        r += 1

        _write_header(ws, r, ["Tanggal", "Keterangan", "Ref", "Debet", "Kredit", "Saldo"])
        r += 1

        for row in item["rows"]:
            ws.cell(row=r, column=1, value=row["tanggal"].strftime("%d/%m/%Y"))
            ws.cell(row=r, column=2, value=row["keterangan"])
            # Untuk Excel, semua Ref dipaksa "JU1" sesuai permintaan
            ws.cell(row=r, column=3, value="JU1").alignment = _ALIGN_CENTER
            for col, val in [(4, row["debet"]), (5, row["kredit"]), (6, row["saldo"])]:
                cell = ws.cell(row=r, column=col, value=val if (col == 6 or val) else None)
                cell.number_format = _NUM_FMT
                cell.alignment = _ALIGN_RIGHT
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = _BORDER_THIN
            r += 1

        # Saldo akhir
        ws.cell(row=r, column=1, value="Saldo Akhir").font = _FONT_BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=6, value=item["saldo_akhir"])
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)
        r += 2  # spacer

    return _as_bytes(wb)


# ── 3. Neraca Saldo ──────────────────────────────────────────────────────────
def build_neraca_saldo(company, period_str, rows, sum_debet, sum_kredit, seimbang) -> bytes:
    """rows: list of {akun, debet, kredit}"""
    wb = Workbook(); ws = wb.active; ws.title = "Neraca Saldo"
    _set_widths(ws, [10, 40, 18, 18])
    r = _letterhead(ws, company, "Neraca Saldo", period_str, 4)

    _write_header(ws, r, ["Kode", "Nama Akun", "Debet", "Kredit"])
    r += 1

    for row in rows:
        ws.cell(row=r, column=1, value=row["akun"].kode_akun)
        ws.cell(row=r, column=2, value=row["akun"].nama_akun)
        for col, val in [(3, row["debet"]), (4, row["kredit"])]:
            c = ws.cell(row=r, column=col, value=val if val else None)
            c.number_format = _NUM_FMT
            c.alignment = _ALIGN_RIGHT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = _BORDER_THIN
        r += 1

    # Total
    ws.cell(row=r, column=1, value="Total").font = _FONT_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for col, val in [(3, sum_debet), (4, sum_kredit)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)
    r += 1

    # Status keseimbangan
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    status = "✓ Neraca saldo seimbang" if seimbang else f"✗ Tidak seimbang — selisih {abs(sum_debet - sum_kredit):,.0f}"
    c = ws.cell(row=r, column=1, value=status)
    c.alignment = _ALIGN_CENTER
    c.font = Font(name="Arial", size=10, bold=True,
                  color="1E8449" if seimbang else "C0392B")

    return _as_bytes(wb)


# ── 4. Worksheet (Kertas Kerja) ──────────────────────────────────────────────
def build_worksheet(company, period_str, rows, totals, selisih, grand) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Kertas Kerja"
    _set_widths(ws, [28] + [12] * 10)
    r = _letterhead(ws, company, "Kertas Kerja (Neraca Lajur)", period_str, 11)

    # Header dua baris: section name + Debet/Kredit
    sections = ["Neraca Saldo", "AJP", "NSD", "Laba Rugi", "Neraca"]
    ws.cell(row=r, column=1, value="Keterangan").font = _FONT_HEADER
    ws.cell(row=r, column=1).alignment = _ALIGN_CENTER
    ws.cell(row=r, column=1).fill = _FILL_HEADER
    ws.cell(row=r, column=1).border = _BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)

    for i, name in enumerate(sections):
        col = 2 + i * 2
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        c = ws.cell(row=r, column=col, value=name)
        c.font = _FONT_HEADER; c.alignment = _ALIGN_CENTER
        c.fill = _FILL_HEADER; c.border = _BORDER_THIN
        ws.cell(row=r, column=col + 1).border = _BORDER_THIN
        ws.cell(row=r, column=col + 1).fill = _FILL_HEADER

    # Sub-header (Debet / Kredit)
    for i in range(5):
        col = 2 + i * 2
        for offset, lbl in enumerate(("Debet", "Kredit")):
            c = ws.cell(row=r + 1, column=col + offset, value=lbl)
            c.font = _FONT_HEADER; c.alignment = _ALIGN_CENTER
            c.fill = _FILL_HEADER; c.border = _BORDER_THIN

    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 18
    r += 2

    # Body
    keys = ["ns_d","ns_k","ajp_d","ajp_k","nsd_d","nsd_k","lr_d","lr_k","n_d","n_k"]
    for row in rows:
        ws.cell(row=r, column=1, value=f"{row['kode_akun']}  {row['nama_akun']}")
        for i, k in enumerate(keys):
            v = row[k]
            c = ws.cell(row=r, column=2 + i, value=v if v else None)
            c.number_format = _NUM_FMT
            c.alignment = _ALIGN_RIGHT
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = _BORDER_THIN
        r += 1

    # Total
    ws.cell(row=r, column=1, value="Total").font = _FONT_BOLD
    for i, k in enumerate(keys):
        c = ws.cell(row=r, column=2 + i, value=totals[k] if totals[k] else None)
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
    ws.cell(row=r, column=1).fill = _FILL_TOTAL
    for c in range(1, 12):
        ws.cell(row=r, column=c).border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)
    r += 1

    # Selisih (Laba / Rugi)
    if selisih.get("label"):
        ws.cell(row=r, column=1, value=selisih["label"]).font = _FONT_BOLD
        for col_idx, key in [(8, "lr_d"), (9, "lr_k"), (10, "n_d"), (11, "n_k")]:
            v = selisih[key]
            c = ws.cell(row=r, column=col_idx, value=v if v else None)
            c.font = _FONT_BOLD; c.number_format = _NUM_FMT
            c.alignment = _ALIGN_RIGHT
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = _BORDER_THIN
        r += 1

    # Grand Total
    ws.cell(row=r, column=1, value="Grand Total").font = _FONT_BOLD
    for col_idx, key in [(8, "lr_d"), (9, "lr_k"), (10, "n_d"), (11, "n_k")]:
        c = ws.cell(row=r, column=col_idx, value=grand[key] if grand[key] else None)
        c.font = _FONT_BOLD; c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT; c.fill = _FILL_TOTAL
    ws.cell(row=r, column=1).fill = _FILL_TOTAL
    for c in range(1, 12):
        ws.cell(row=r, column=c).border = Border(top=_MED, bottom=_MED, left=_THIN, right=_THIN)

    # Freeze panes: kolom keterangan + 2 baris header sticky (data mulai row 7)
    ws.freeze_panes = "B7"
    return _as_bytes(wb)


# ── 5. Laba Rugi ─────────────────────────────────────────────────────────────
def build_laba_rugi(company, period_str, lr) -> bytes:
    """lr: {pendapatan_items, beban_items, total_pendapatan, total_beban, net, is_laba}"""
    wb = Workbook(); ws = wb.active; ws.title = "Laba Rugi"
    _set_widths(ws, [40, 18, 18])
    r = _letterhead(ws, company, "Laporan Laba Rugi", period_str, 3)

    def write_money(row, col, val, bold=False, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = _NUM_FMT
        c.alignment = _ALIGN_RIGHT
        if bold: c.font = _FONT_BOLD
        if fill: c.fill = fill

    # Pendapatan
    c = ws.cell(row=r, column=1, value="Pendapatan Diterima"); c.font = _FONT_BOLD
    r += 1
    for akun, bal in lr["pendapatan_items"]:
        ws.cell(row=r, column=1, value=akun.nama_akun).alignment = Alignment(horizontal="left", indent=2)
        write_money(r, 2, bal)
        r += 1
    ws.cell(row=r, column=1, value="Total Pendapatan").font = _FONT_BOLD
    write_money(r, 2, lr["total_pendapatan"], bold=True)
    ws.cell(row=r, column=2).border = Border(top=_THIN)
    r += 2

    # Beban
    c = ws.cell(row=r, column=1, value="Beban Operasional"); c.font = _FONT_BOLD
    r += 1
    for akun, bal in lr["beban_items"]:
        ws.cell(row=r, column=1, value=akun.nama_akun).alignment = Alignment(horizontal="left", indent=2)
        write_money(r, 2, bal)
        r += 1
    ws.cell(row=r, column=1, value="Total Beban").font = _FONT_BOLD
    write_money(r, 2, lr["total_beban"], bold=True)
    ws.cell(row=r, column=2).border = Border(top=_THIN)
    r += 2

    # Net
    label = "Net Income" if lr["is_laba"] else "Net Loss"
    ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
    write_money(r, 3, lr["net"], bold=True)
    ws.cell(row=r, column=3).border = Border(top=_MED, bottom=Side(style="double", color="111111"))

    return _as_bytes(wb)


# ── 6. Ekuitas Pemilik ───────────────────────────────────────────────────────
_BULAN = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
          'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']


def _fmt_tgl(d):
    return f"{d.day} {_BULAN[d.month]} {d.year}"


def build_ekuitas(company, period_str, ek) -> bytes:
    """ek: {nama_modal, modal_awal, tgl_awal, tgl_akhir, add_items, less_items, perubahan, modal_akhir}"""
    wb = Workbook(); ws = wb.active; ws.title = "Ekuitas Pemilik"
    _set_widths(ws, [40, 18, 18])
    r = _letterhead(ws, company, "Laporan Ekuitas Pemilik", period_str, 3)

    # Modal Awal
    ws.cell(row=r, column=1, value=f"{ek['nama_modal']}, {_fmt_tgl(ek['tgl_awal'])}")
    c = ws.cell(row=r, column=3, value=ek["modal_awal"])
    c.number_format = _NUM_FMT; c.alignment = _ALIGN_RIGHT
    r += 1

    # Add items
    for i, (label, amt) in enumerate(ek["add_items"]):
        if i == 0:
            ws.cell(row=r, column=1, value=f"Add:  {label}").alignment = Alignment(horizontal="left", indent=1)
            ws.cell(row=r, column=1).font = Font(name="Arial", size=10, italic=True, color="1A6B1A")
        else:
            ws.cell(row=r, column=1, value=label).alignment = Alignment(horizontal="left", indent=3)
        c = ws.cell(row=r, column=2, value=amt)
        c.number_format = _NUM_FMT; c.alignment = _ALIGN_RIGHT
        r += 1

    # Less items
    for i, (label, amt) in enumerate(ek["less_items"]):
        if i == 0:
            ws.cell(row=r, column=1, value=f"Less: {label}").alignment = Alignment(horizontal="left", indent=1)
            ws.cell(row=r, column=1).font = Font(name="Arial", size=10, italic=True, color="8B1A1A")
        else:
            ws.cell(row=r, column=1, value=label).alignment = Alignment(horizontal="left", indent=3)
        c = ws.cell(row=r, column=2, value=amt)
        c.number_format = _NUM_FMT; c.alignment = _ALIGN_RIGHT
        r += 1

    # Perubahan
    c = ws.cell(row=r, column=3, value=ek["perubahan"])
    c.number_format = _NUM_FMT; c.alignment = _ALIGN_RIGHT
    c.border = Border(top=_THIN)
    r += 1

    # Modal akhir
    ws.cell(row=r, column=1, value=f"{ek['nama_modal']}, {_fmt_tgl(ek['tgl_akhir'])}").font = _FONT_BOLD
    c = ws.cell(row=r, column=3, value=ek["modal_akhir"])
    c.font = _FONT_BOLD; c.number_format = _NUM_FMT; c.alignment = _ALIGN_RIGHT
    c.border = Border(top=_MED, bottom=Side(style="double", color="111111"))

    return _as_bytes(wb)


# ── 7. Neraca ────────────────────────────────────────────────────────────────
def build_neraca(company, period_str, nr) -> bytes:
    """nr: {rows: [((l_lbl,l_amt,l_type),(r_lbl,r_amt,r_type))], total_aset, total_km, seimbang}"""
    wb = Workbook(); ws = wb.active; ws.title = "Neraca"
    _set_widths(ws, [30, 16, 30, 16])
    r = _letterhead(ws, company, "Neraca", period_str, 4)

    # Sub-header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    for col, label in [(1, "ASET"), (3, "KEWAJIBAN & MODAL")]:
        c = ws.cell(row=r, column=col, value=label)
        c.font = _FONT_HEADER; c.alignment = _ALIGN_CENTER
        c.fill = _FILL_HEADER
        c.border = _BORDER_THIN
        # Borders on merged cell partner
        ws.cell(row=r, column=col + 1).border = _BORDER_THIN
        ws.cell(row=r, column=col + 1).fill = _FILL_HEADER
    ws.row_dimensions[r].height = 20
    r += 1

    def _write_side(row, col_lbl, col_amt, lbl, amt, ttype):
        cl = ws.cell(row=row, column=col_lbl, value=lbl)
        ca = ws.cell(row=row, column=col_amt)

        if ttype == "section":
            cl.font = _FONT_BOLD
        elif ttype == "item":
            cl.alignment = Alignment(horizontal="left", indent=2)
        elif ttype == "item-kontra":
            cl.alignment = Alignment(horizontal="left", indent=2)
            cl.font = Font(name="Arial", size=10, color="666666")
        elif ttype == "subtotal":
            cl.font = _FONT_BOLD
            ca.border = Border(top=_THIN)
        elif ttype == "total":
            cl.font = _FONT_BOLD
            cl.fill = _FILL_TOTAL
            ca.fill = _FILL_TOTAL
            ca.border = Border(top=_MED, bottom=Side(style="double", color="111111"))

        if amt is not None:
            if ttype == "item-kontra":
                ca.value = -amt
                ca.number_format = '(#,##0);(#,##0)'
                ca.font = Font(name="Arial", size=10, color="666666")
            else:
                ca.value = amt
                ca.number_format = _NUM_FMT
            ca.alignment = _ALIGN_RIGHT
            if ttype in ("subtotal", "total"):
                ca.font = _FONT_BOLD if ttype != "item-kontra" else ca.font

    for (l_lbl, l_amt, l_type), (r_lbl, r_amt, r_type) in nr["rows"]:
        _write_side(r, 1, 2, l_lbl, l_amt, l_type)
        _write_side(r, 3, 4, r_lbl, r_amt, r_type)
        r += 1

    if not nr.get("seimbang", True):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value="⚠ Neraca tidak seimbang!")
        c.font = Font(name="Arial", size=10, bold=True, color="C0392B")
        c.alignment = _ALIGN_CENTER

    return _as_bytes(wb)
